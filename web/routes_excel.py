import json
from datetime import datetime
from pathlib import Path
from tempfile import NamedTemporaryFile

from fastapi import APIRouter, HTTPException, Query, UploadFile
from fastapi.responses import JSONResponse
from openpyxl import load_workbook
from pydantic import BaseModel

from storage.excel import ExcelCaseRepository
from web.routes_config import _load_yaml, _save_yaml, _get_input_path
from web.files import INPUTS_DIR, get_existing_input_path, project_relative, resolve_config_input_path

router = APIRouter(prefix="/api/excel", tags=["excel"])

MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB
SETS_META_FILE = INPUTS_DIR / ".sets_meta.json"


class SetMetaRequest(BaseModel):
    """测试集文件级元数据。"""

    name: str | None = None
    description: str = ""


def _normalize_description(value: str | None) -> str:
    """规范化测试集说明，避免本地元数据无限膨胀。"""
    return (value or "").strip()[:1000]


def _normalize_set_name(value: str | None) -> str:
    """规范化测试集展示名称。"""
    return (value or "").strip()[:120]


def _default_set_name(filename: str) -> str:
    """测试集默认展示名称：原始文件名去后缀。"""
    return Path(filename).stem


def _display_name_for(filename: str, meta: dict[str, dict]) -> str:
    """获取文件当前展示名称。"""
    return _normalize_set_name(str(meta.get(filename, {}).get("name") or "")) or _default_set_name(filename)


def _ensure_unique_set_name(filename: str, name: str, meta: dict[str, dict]) -> None:
    """确保测试集展示名称在所有测试集内唯一。"""
    normalized = _normalize_set_name(name)
    for file_info in _scan_input_files():
        other_filename = file_info["filename"]
        if other_filename == filename:
            continue
        if _display_name_for(other_filename, meta) == normalized:
            raise HTTPException(400, f"名称已存在: {normalized}")


def _read_sets_meta_file() -> dict:
    """读取测试集文件级元数据。"""
    if not SETS_META_FILE.is_file():
        return {}
    try:
        with open(SETS_META_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except (json.JSONDecodeError, OSError):
        return {}


def _load_sets_meta() -> dict[str, dict]:
    """读取存在测试集的元数据，过滤过期文件记录。"""
    data = _read_sets_meta_file()
    result: dict[str, dict] = {}
    for filename, meta in data.items():
        if not isinstance(meta, dict):
            continue
        try:
            path = get_existing_input_path(filename)
        except HTTPException:
            continue
        description = _normalize_description(str(meta.get("description") or ""))
        name = _normalize_set_name(str(meta.get("name") or ""))
        item: dict[str, str] = {}
        if name:
            item["name"] = name
        if description:
            item["description"] = description
        if item:
            result[path.name] = item
    return result


def _save_sets_meta(meta: dict[str, dict]) -> None:
    """保存测试集文件级元数据。"""
    INPUTS_DIR.mkdir(parents=True, exist_ok=True)
    data: dict[str, dict[str, str]] = {}
    for filename, item in sorted(meta.items()):
        name = _normalize_set_name(str(item.get("name") or ""))
        description = _normalize_description(str(item.get("description") or ""))
        saved: dict[str, str] = {}
        if name:
            saved["name"] = name
        if description:
            saved["description"] = description
        if saved:
            data[filename] = saved
    with open(SETS_META_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _remove_set_meta(filename: str) -> None:
    """删除某个测试集对应的文件级元数据。"""
    meta = _load_sets_meta()
    meta.pop(filename, None)
    _save_sets_meta(meta)


def _scan_input_files() -> list[dict]:
    """扫描 ``inputs/`` 目录下的所有 Excel 文件。

    Returns:
        文件信息列表，按修改时间倒序排列。
    """
    if not INPUTS_DIR.is_dir():
        return []
    files: list[dict] = []
    for entry in INPUTS_DIR.iterdir():
        if entry.is_file() and entry.suffix.lower() in (".xlsx", ".xlsm"):
            stat = entry.stat()
            files.append(
                {
                    "filename": entry.name,
                    "size": stat.st_size,
                    "updated_at": datetime.fromtimestamp(stat.st_mtime).strftime(
                        "%Y-%m-%dT%H:%M:%S"
                    ),
                }
            )
    files.sort(key=lambda f: f["updated_at"], reverse=True)
    return files


def _detect_sheets(file_path: Path) -> list[dict]:
    """读取 Excel 文件中所有 sheet 的名称和实际用例数。

    Args:
        file_path: Excel 文件路径。

    Returns:
        sheet 信息列表，每项包含 ``name`` 和 ``rows``（真实用例数）。
    """
    wb = load_workbook(file_path, read_only=True)
    sheet_names = list(wb.sheetnames)
    wb.close()

    sheets: list[dict] = []
    for name in sheet_names:
        try:
            repo = ExcelCaseRepository(file_path, name)
            cases = repo.read_cases()
        except ValueError as exc:
            raise HTTPException(400, str(exc)) from exc
        sheets.append({"name": name, "rows": len(cases)})
    return sheets


def _set_current_excel(config: dict, path: Path, sheets: list[dict] | None = None) -> None:
    """把当前配置切换到指定 Excel，并保持 sheet 有效。"""
    if sheets is None:
        sheets = _detect_sheets(path)
    sheet_names = [sheet["name"] for sheet in sheets]
    excel_cfg = config.setdefault("excel", {})
    current_sheet = excel_cfg.get("sheet_name")
    excel_cfg["input_path"] = project_relative(path)
    excel_cfg["sheet_name"] = (
        current_sheet if current_sheet in sheet_names else sheet_names[0]
    )


@router.post("/upload")
async def upload_excel(file: UploadFile) -> JSONResponse:
    """上传测试集 Excel 文件，同名文件覆盖。

    Args:
        file: 上传的 .xlsx 或 .xlsm 文件。

    Returns:
        文件名及所有 sheet 的名称和行数。

    Raises:
        HTTPException 400: 文件类型或大小不符合要求。
    """
    filename = _get_input_path(file.filename).name

    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(400, f"文件大小不能超过 {MAX_FILE_SIZE // 1024 // 1024} MB")

    INPUTS_DIR.mkdir(parents=True, exist_ok=True)
    dest = _get_input_path(filename)
    tmp_path: Path | None = None
    try:
        with NamedTemporaryFile(
            mode="wb",
            delete=False,
            dir=INPUTS_DIR,
            suffix=dest.suffix,
        ) as tmp:
            tmp.write(content)
            tmp_path = Path(tmp.name)

        sheets = _detect_sheets(tmp_path)
        tmp_path.replace(dest)
    except HTTPException:
        if tmp_path is not None:
            tmp_path.unlink(missing_ok=True)
        raise
    except Exception as exc:
        if tmp_path is not None:
            tmp_path.unlink(missing_ok=True)
        raise HTTPException(400, f"Excel 文件不可读取: {exc}") from exc

    # 自动设置为当前使用的文件，sheet 保持当前配置或默认
    config = _load_yaml()
    _set_current_excel(config, dest, sheets)
    _save_yaml(config)

    return JSONResponse({"filename": filename, "sheets": sheets})


@router.get("/sets")
def list_sets(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    sort_by: str = Query("updated_at"),
    sort_dir: str = Query("desc"),
    name_query: str = Query(""),
) -> JSONResponse:
    """列出 ``inputs/`` 下所有已上传的 Excel 测试集（分页）。

    Args:
        page: 页码，从 1 开始。
        page_size: 每页条数，默认 20，上限 200。

    Returns:
        当前页文件列表、分页信息和 inputs 目录路径。
    """
    if sort_by not in {"updated_at"}:
        raise HTTPException(400, "sort_by 仅支持 updated_at")
    reverse = sort_dir.lower() != "asc"

    query = name_query.strip().casefold()
    meta_data = _load_sets_meta()
    all_files = _scan_input_files()
    for f in all_files:
        file_meta = meta_data.get(f["filename"], {})
        f["name"] = file_meta.get("name") or _default_set_name(f["filename"])
        f["description"] = file_meta.get("description", "")

    if query:
        all_files = [
            f for f in all_files if query in str(f.get("name", "")).casefold()
        ]

    all_files.sort(key=lambda f: f["updated_at"], reverse=reverse)
    total = len(all_files)

    # 分页切片
    start = (page - 1) * page_size
    end = start + page_size
    page_files = all_files[start:end]

    config = _load_yaml()
    excel_cfg = config.get("excel", {})
    current_path = excel_cfg.get("input_path", "")
    current = ""
    if current_path:
        try:
            current = resolve_config_input_path(current_path).name
        except HTTPException:
            current = ""
    return JSONResponse({
        "files": page_files,
        "current": current,
        "inputs_dir": str(INPUTS_DIR.resolve()),
        "total": total,
        "page": page,
        "page_size": page_size,
    })


@router.get("/sets/{filename}/meta")
def get_set_meta(filename: str) -> JSONResponse:
    """读取单个测试集的文件级元数据。"""
    path = get_existing_input_path(filename)
    meta = _load_sets_meta().get(path.name, {})
    return JSONResponse(
        {
            "filename": path.name,
            "name": meta.get("name") or _default_set_name(path.name),
            "description": meta.get("description", ""),
        }
    )


@router.put("/sets/{filename}/meta")
def update_set_meta(filename: str, body: SetMetaRequest) -> JSONResponse:
    """更新单个测试集的文件级元数据。"""
    path = get_existing_input_path(filename)
    meta = _load_sets_meta()
    item = dict(meta.get(path.name, {}))

    if "name" in body.model_fields_set:
        name = _normalize_set_name(body.name)
        if not name:
            raise HTTPException(400, "名称不能为空")
        _ensure_unique_set_name(path.name, name, meta)
        if name and name != path.stem:
            item["name"] = name
        else:
            item.pop("name", None)

    if "description" in body.model_fields_set:
        description = _normalize_description(body.description)
        if description:
            item["description"] = description
        else:
            item.pop("description", None)

    if item:
        meta[path.name] = item
    else:
        meta.pop(path.name, None)
    _save_sets_meta(meta)
    return JSONResponse(
        {
            "filename": path.name,
            "name": item.get("name") or path.stem,
            "description": item.get("description", ""),
        }
    )


@router.get("/sheets")
def list_sheets(filename: str | None = None) -> JSONResponse:
    """列出指定或当前 Excel 文件的所有 sheet。

    Args:
        filename: Excel 文件名。不传则使用当前配置的文件。

    Returns:
        sheet 名称和行数列表。

    Raises:
        HTTPException 400: 文件不存在。
    """
    if filename:
        input_path = get_existing_input_path(filename)
    else:
        config = _load_yaml()
        excel_cfg = config.get("excel", {})
        input_path_str = excel_cfg.get("input_path", "inputs/testcases.xlsx")
        input_path = resolve_config_input_path(input_path_str)

    if not input_path.is_file():
        raise HTTPException(400, f"文件不存在: {input_path.name}")

    sheets = _detect_sheets(input_path)
    return JSONResponse({"filename": input_path.name, "sheets": sheets})


@router.get("/refresh")
def refresh_cases() -> JSONResponse:
    """重新加载当前 Excel + sheet 的用例数据。

    用于用户在外部修改 Excel 后刷新。

    Returns:
        最新的用例列表和元信息。

    Raises:
        HTTPException 400: 配置的 Excel 文件不存在。
    """
    config = _load_yaml()
    excel_cfg = config.get("excel", {})
    input_path_str = excel_cfg.get("input_path", "inputs/testcases.xlsx")
    input_path = resolve_config_input_path(input_path_str)
    if not input_path.is_file():
        raise HTTPException(400, f"Excel 文件不存在: {input_path}")

    sheet_name = excel_cfg.get("sheet_name", "Sheet1")
    try:
        repo = ExcelCaseRepository(input_path, sheet_name)
        raw_cases = repo.read_cases()
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc

    cases = [{"case_id": c.case_id, "question": c.question} for c in raw_cases]
    return JSONResponse(
        {
            "cases": cases,
            "count": len(cases),
            "filename": input_path.name,
            "sheet_name": sheet_name,
            "refreshed_at": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
        }
    )


@router.delete("/sets/{filename}")
def delete_set(filename: str) -> JSONResponse:
    """删除一个测试集 Excel 文件。

    Args:
        filename: 要删除的文件名。

    Returns:
        确认信息。

    Raises:
        HTTPException 404: 文件不存在。
    """
    input_path = _get_input_path(filename)
    if not input_path.is_file():
        raise HTTPException(404, f"文件不存在: {filename}")

    input_path.unlink()
    _remove_set_meta(input_path.name)

    config = _load_yaml()
    excel_cfg = config.setdefault("excel", {})
    try:
        current_path = resolve_config_input_path(excel_cfg.get("input_path", ""))
    except HTTPException:
        current_path = None
    if current_path == input_path.resolve():
        remaining = _scan_input_files()
        if remaining:
            next_path = _get_input_path(remaining[0]["filename"])
            _set_current_excel(config, next_path)
        else:
            excel_cfg["input_path"] = "inputs/testcases.xlsx"
            excel_cfg["sheet_name"] = "Sheet1"
        _save_yaml(config)

    return JSONResponse({"ok": True, "filename": input_path.name})
