"""首 Sheet 读取、请求模板渲染和 Run 输入快照。"""

from __future__ import annotations

import hashlib
import json
from copy import deepcopy
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from execution.models import CaseRunRecord, RunRecord, TargetRecord
from execution.repository import RunRepository
from storage.excel import TestCase, read_test_cases_from_sheet


class RunPreparationError(ValueError):
    """测试集或请求模板不能生成合法运行输入。"""


@dataclass(frozen=True)
class ExcelExecutionSnapshot:
    """从工作簿同一字节版本读取的首 Sheet 快照。"""

    filename: str
    sha256: str
    sheet_name: str
    sheet_count: int
    ignored_sheet_names: tuple[str, ...]
    cases: tuple[TestCase, ...]


@dataclass(frozen=True)
class PreparedCase:
    """一条待执行 CaseRun 及其已渲染请求体。"""

    record: CaseRunRecord
    request_body: Any


@dataclass(frozen=True)
class PreparedRun:
    """可原子持久化的 Run 输入。"""

    record: RunRecord
    cases: tuple[PreparedCase, ...]


def _reject_nonstandard_constant(value: str) -> None:
    raise RunPreparationError(f"请求模板包含非标准 JSON 数值: {value}")


def _unique_object(pairs: list[tuple[str, Any]]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, value in pairs:
        if key in result:
            raise RunPreparationError(f"请求模板包含重复字段: {key}")
        result[key] = value
    return result


def parse_request_template(source: str) -> Any:
    """严格解析 JSON 请求模板，不接受 NaN、Infinity 或重复字段。"""
    if not isinstance(source, str) or not source.strip():
        raise RunPreparationError("请求模板必须是非空 JSON")
    try:
        return json.loads(
            source,
            parse_constant=_reject_nonstandard_constant,
            object_pairs_hook=_unique_object,
        )
    except RunPreparationError:
        raise
    except (json.JSONDecodeError, RecursionError) as exc:
        raise RunPreparationError(f"请求模板不是合法 JSON: {exc}") from exc


def normalize_request_template(template: Any) -> Any:
    """复制并验证已经结构化的 JSON 模板。"""
    try:
        serialized = json.dumps(template, ensure_ascii=False, allow_nan=False)
    except (TypeError, ValueError, RecursionError) as exc:
        raise RunPreparationError(f"请求模板不是合法 JSON: {exc}") from exc
    return parse_request_template(serialized)


def render_request_body(template: Any, question: str) -> Any:
    """只递归替换 JSON 字符串值中的 question 变量。"""
    if isinstance(template, str):
        return template.replace("${question}", question)
    if isinstance(template, list):
        return [render_request_body(item, question) for item in template]
    if isinstance(template, dict):
        return {
            key: render_request_body(value, question)
            for key, value in template.items()
        }
    return deepcopy(template)


def read_first_sheet_snapshot(file_path: str | Path) -> ExcelExecutionSnapshot:
    """从同一份 Excel 字节内容计算哈希并读取首个 Sheet。"""
    path = Path(file_path)
    if not path.is_file():
        raise RunPreparationError(f"Excel 文件不存在: {path}")
    try:
        content = path.read_bytes()
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
        raise RunPreparationError(f"Excel 文件不可读取: {exc}") from exc
    try:
        sheet_names = tuple(workbook.sheetnames)
        if not sheet_names:
            raise RunPreparationError("Excel 工作簿不包含 Sheet")
        cases = tuple(read_test_cases_from_sheet(workbook[sheet_names[0]]))
    finally:
        workbook.close()
    return ExcelExecutionSnapshot(
        filename=path.name,
        sha256=hashlib.sha256(content).hexdigest(),
        sheet_name=sheet_names[0],
        sheet_count=len(sheet_names),
        ignored_sheet_names=sheet_names[1:],
        cases=cases,
    )


class RunPreparationService:
    """构建并持久化不受后续配置变化影响的 Run 输入。"""

    SNAPSHOT_SCHEMA_VERSION = 1

    def __init__(self, repository: RunRepository):
        self.repository = repository

    def prepare(
        self,
        *,
        testset_path: str | Path,
        request_template: Any,
        target: TargetRecord | None = None,
        workflow_id: str | None = None,
        workflow_snapshot: dict[str, Any] | None = None,
        parameters: dict[str, Any] | None = None,
        run_id: str | None = None,
    ) -> PreparedRun:
        excel = read_first_sheet_snapshot(testset_path)
        template = (
            parse_request_template(request_template)
            if isinstance(request_template, str)
            else normalize_request_template(request_template)
        )
        snapshot: dict[str, Any] = {
            "schema_version": self.SNAPSHOT_SCHEMA_VERSION,
            "excel": {
                "filename": excel.filename,
                "sha256": excel.sha256,
                "sheet_name": excel.sheet_name,
                "sheet_count": excel.sheet_count,
                "ignored_sheet_names": list(excel.ignored_sheet_names),
            },
            "request_template": template,
        }
        if target is not None:
            snapshot["target"] = target.model_dump(mode="json")
        if workflow_snapshot is not None:
            snapshot["workflow"] = normalize_request_template(workflow_snapshot)

        run_kwargs: dict[str, Any] = {
            "testset_filename": excel.filename,
            "sheet_name": excel.sheet_name,
            "target_id": target.id if target else None,
            "workflow_id": workflow_id,
            "parameters": normalize_request_template(parameters or {}),
            "snapshot": snapshot,
        }
        if run_id is not None:
            run_kwargs["id"] = run_id
        run = RunRecord(**run_kwargs)
        cases = tuple(
            PreparedCase(
                record=CaseRunRecord(
                    run_id=run.id,
                    case_id=case.case_id,
                    row_number=case.row_number,
                    question=case.question,
                ),
                request_body=render_request_body(template, case.question),
            )
            for case in excel.cases
        )
        return PreparedRun(record=run, cases=cases)

    def create(self, **kwargs: Any) -> PreparedRun:
        """构建快照并原子创建 Run 与全部 CaseRun。"""
        prepared = self.prepare(**kwargs)
        self.repository.create_run_with_cases(
            prepared.record,
            tuple(case.record for case in prepared.cases),
        )
        return prepared
