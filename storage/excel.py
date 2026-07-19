from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


CASE_ID_HEADERS = {"case_id", "case id", "用例id", "用例编号"}
QUESTION_HEADERS = {"question", "问题"}


def _as_text(value) -> str:
    """把 Excel 单元格值转换为去除首尾空白的文本。"""
    return "" if value is None else str(value).strip()


@dataclass(frozen=True)
class TestCase:
    """从 Excel 读取的一条 Web 管理用测试用例。"""

    case_id: str
    question: str
    row_number: int


def read_test_cases_from_sheet(sheet) -> list[TestCase]:
    """从一个已打开的工作表读取有效用例并保留 Excel 行号。"""
    rows = sheet.iter_rows(min_col=1, max_col=2, values_only=True)
    testcases: list[TestCase] = []
    seen_ids: set[str] = set()
    for row_number, row in enumerate(rows, start=1):
        values = list(row) + [None] * (2 - len(row))
        case_id = _as_text(values[0])
        question = _as_text(values[1])

        if row_number == 1 and (
            case_id.casefold() in CASE_ID_HEADERS
            or question.casefold() in QUESTION_HEADERS
        ):
            continue
        if not case_id or not question:
            continue
        if case_id in seen_ids:
            continue
        seen_ids.add(case_id)
        testcases.append(
            TestCase(
                case_id=case_id,
                question=question,
                row_number=row_number,
            )
        )
    return testcases


class ExcelCaseRepository:
    """读取测试用例 Excel。

    当前 Web 项目只支持固定两列输入格式：``case_id | question``。
    第三列及之后允许存在历史结果或人工备注，但不会参与用例读取。
    """

    def __init__(self, file_path: str | Path, sheet_name: str = "Sheet1"):
        """绑定一个测试用例工作簿及目标工作表。"""
        self.path = Path(file_path)
        self.sheet_name = sheet_name

    def _select_sheet(self, workbook):
        """选择指定工作表。"""
        if self.sheet_name in workbook.sheetnames:
            return workbook[self.sheet_name]
        available = ", ".join(workbook.sheetnames)
        raise ValueError(f"Sheet 不存在: {self.sheet_name}。可用: {available}")

    def read_cases(self) -> list[TestCase]:
        """读取工作簿中的有效测试用例。

        空行、表头、空 ID、空问题和重复 ID 会被忽略。
        """
        if not self.path.is_file():
            raise FileNotFoundError(f"Excel 文件不存在: {self.path}")

        workbook = load_workbook(self.path, read_only=True, data_only=True)
        try:
            sheet = self._select_sheet(workbook)
            return read_test_cases_from_sheet(sheet)
        finally:
            workbook.close()
