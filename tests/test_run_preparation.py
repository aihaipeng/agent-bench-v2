import hashlib
import json

import pytest
from openpyxl import Workbook

from execution import (
    CaseRunRecord,
    RunPreparationError,
    RunPreparationService,
    RunRecord,
    RunRepository,
    RunRepositoryError,
    TargetRecord,
    normalize_request_template,
    parse_request_template,
    read_first_sheet_snapshot,
    render_request_body,
)


def _save_workbook(path, *, first_question='他说 "你好"\n下一行') -> None:
    workbook = Workbook()
    first = workbook.active
    first.title = "实际首个 Sheet"
    first.append(["case_id", "question", "旧结果"])
    first.append(["case_001", first_question, "PASS"])
    first.append(["", "缺少 ID", "FAIL"])
    first.append(["case_001", "重复 ID", "FAIL"])
    first.append(["case_002", "第二个中文问题", "PASS"])
    ignored = workbook.create_sheet("配置中曾选择的 Sheet")
    ignored.append(["case_id", "question"])
    ignored.append(["ignored_001", "不应执行"])
    workbook.create_sheet("第三个 Sheet")
    workbook.save(path)


def _target() -> TargetRecord:
    return TargetRecord(
        id="target-1",
        name="企业 Agent 内网",
        base_url="http://127.0.0.1:9000",
        path="/api/agent/invoke",
        headers={"Authorization": "plain-token"},
        target_total_concurrency=4,
    )


def test_first_sheet_snapshot_uses_one_byte_version_and_preserves_rows(tmp_path):
    workbook_path = tmp_path / "cases.xlsx"
    _save_workbook(workbook_path)
    expected_hash = hashlib.sha256(workbook_path.read_bytes()).hexdigest()

    snapshot = read_first_sheet_snapshot(workbook_path)

    assert snapshot.filename == "cases.xlsx"
    assert snapshot.sha256 == expected_hash
    assert snapshot.sheet_name == "实际首个 Sheet"
    assert snapshot.sheet_count == 3
    assert snapshot.ignored_sheet_names == (
        "配置中曾选择的 Sheet",
        "第三个 Sheet",
    )
    assert [case.case_id for case in snapshot.cases] == ["case_001", "case_002"]
    assert [case.row_number for case in snapshot.cases] == [2, 5]
    assert snapshot.cases[0].question == '他说 "你好"\n下一行'


def test_request_template_recursively_replaces_values_without_touching_keys():
    template = {
        "question": "前缀 ${question} 后缀 ${question}",
        "nested": [
            {"message": "${question}"},
            42,
            True,
            None,
        ],
        "${question}": "键名不替换",
        "username": "tester",
    }
    original = json.loads(json.dumps(template, ensure_ascii=False))
    question = '包含 "引号"、\n换行和中文'

    rendered = render_request_body(template, question)

    assert rendered == {
        "question": f"前缀 {question} 后缀 {question}",
        "nested": [{"message": question}, 42, True, None],
        "${question}": "键名不替换",
        "username": "tester",
    }
    assert template == original
    assert json.loads(json.dumps(rendered, ensure_ascii=False)) == rendered


def test_template_supports_any_standard_json_and_never_injects_case_id():
    array_template = parse_request_template(
        '["${question}", {"username": "u", "password": "p"}]'
    )
    scalar_template = parse_request_template('"${question}"')

    array_body = render_request_body(array_template, "问题")
    scalar_body = render_request_body(scalar_template, "问题")

    assert array_body == ["问题", {"username": "u", "password": "p"}]
    assert scalar_body == "问题"
    assert "case_id" not in json.dumps(array_body, ensure_ascii=False)


@pytest.mark.parametrize(
    "source, message",
    [
        ("", "非空 JSON"),
        ("{broken", "不是合法 JSON"),
        ('{"value": NaN}', "非标准 JSON"),
        ('{"value": Infinity}', "非标准 JSON"),
        ('{"same": 1, "same": 2}', "重复字段"),
    ],
)
def test_parse_request_template_rejects_ambiguous_or_nonstandard_json(
    source, message
):
    with pytest.raises(RunPreparationError, match=message):
        parse_request_template(source)


def test_normalize_request_template_rejects_non_json_python_values():
    cyclic = {}
    cyclic["self"] = cyclic

    with pytest.raises(RunPreparationError, match="不是合法 JSON"):
        normalize_request_template({"unsupported": {1, 2}})
    with pytest.raises(RunPreparationError, match="不是合法 JSON"):
        normalize_request_template(cyclic)


def test_prepare_freezes_excel_target_template_workflow_and_parameters(tmp_path):
    workbook_path = tmp_path / "cases.xlsx"
    _save_workbook(workbook_path)
    repository = RunRepository(tmp_path / "agent_bench.sqlite3")
    target = _target()
    workflow = {"id": "workflow-1", "steps": [{"tool_id": "tool-1"}]}
    parameters = {"timeout_seconds": 600, "case_concurrency": 2}
    service = RunPreparationService(repository)

    prepared = service.prepare(
        testset_path=workbook_path,
        request_template={
            "question": "${question}",
            "username": "user",
            "password": "plain-secret",
        },
        target=target,
        workflow_id="workflow-1",
        workflow_snapshot=workflow,
        parameters=parameters,
        run_id="run-1",
    )
    original_hash = prepared.record.snapshot["excel"]["sha256"]

    _save_workbook(workbook_path, first_question="文件已变化")
    target.name = "Target 已变化"
    workflow["steps"][0]["tool_id"] = "changed-tool"
    parameters["timeout_seconds"] = 1

    assert prepared.record.id == "run-1"
    assert prepared.record.sheet_name == "实际首个 Sheet"
    assert prepared.record.target_id == "target-1"
    assert prepared.record.workflow_id == "workflow-1"
    assert prepared.record.snapshot["excel"] == {
        "filename": "cases.xlsx",
        "sha256": original_hash,
        "sheet_name": "实际首个 Sheet",
        "sheet_count": 3,
        "ignored_sheet_names": ["配置中曾选择的 Sheet", "第三个 Sheet"],
    }
    assert original_hash != hashlib.sha256(workbook_path.read_bytes()).hexdigest()
    assert prepared.record.snapshot["target"]["name"] == "企业 Agent 内网"
    assert prepared.record.snapshot["target"]["headers"] == {
        "Authorization": "plain-token"
    }
    assert prepared.record.snapshot["workflow"]["steps"][0]["tool_id"] == "tool-1"
    assert prepared.record.parameters == {
        "timeout_seconds": 600,
        "case_concurrency": 2,
    }
    assert [case.record.row_number for case in prepared.cases] == [2, 5]
    assert prepared.cases[0].request_body == {
        "question": '他说 "你好"\n下一行',
        "username": "user",
        "password": "plain-secret",
    }
    assert "case_id" not in prepared.cases[0].request_body


def test_create_persists_run_and_cases_for_restart_without_other_sheets(tmp_path):
    workbook_path = tmp_path / "cases.xlsx"
    _save_workbook(workbook_path)
    database_path = tmp_path / "agent_bench.sqlite3"
    service = RunPreparationService(RunRepository(database_path))

    prepared = service.create(
        testset_path=workbook_path,
        request_template='{"question":"${question}"}',
        target=_target(),
        parameters={"timeout_seconds": 600},
        run_id="persisted-run",
    )

    restarted = RunRepository(database_path)
    restored_run = restarted.get_run("persisted-run")
    restored_cases = restarted.list_case_runs("persisted-run")
    assert restored_run is not None
    assert restored_run.snapshot == prepared.record.snapshot
    assert [(case.case_id, case.row_number, case.question) for case in restored_cases] == [
        ("case_001", 2, '他说 "你好"\n下一行'),
        ("case_002", 5, "第二个中文问题"),
    ]
    assert all(case.case_id != "ignored_001" for case in restored_cases)


def test_run_and_case_bundle_rolls_back_as_one_transaction(tmp_path):
    repository = RunRepository(tmp_path / "agent_bench.sqlite3")
    run = RunRecord(
        id="atomic-run",
        testset_filename="cases.xlsx",
        sheet_name="Sheet1",
    )
    duplicate_cases = (
        CaseRunRecord(
            id="case-run-1",
            run_id=run.id,
            case_id="duplicate",
            row_number=2,
            question="first",
        ),
        CaseRunRecord(
            id="case-run-2",
            run_id=run.id,
            case_id="duplicate",
            row_number=3,
            question="second",
        ),
    )

    with pytest.raises(RunRepositoryError, match="UNIQUE"):
        repository.create_run_with_cases(run, duplicate_cases)

    assert repository.get_run(run.id) is None
    assert repository.list_case_runs(run.id) == []


def test_missing_and_invalid_excel_are_reported_before_run_creation(tmp_path):
    repository = RunRepository(tmp_path / "agent_bench.sqlite3")
    service = RunPreparationService(repository)

    with pytest.raises(RunPreparationError, match="不存在"):
        service.create(
            testset_path=tmp_path / "missing.xlsx",
            request_template="{}",
            run_id="missing-run",
        )
    invalid = tmp_path / "invalid.xlsx"
    invalid.write_bytes(b"not an xlsx file")
    with pytest.raises(RunPreparationError, match="不可读取"):
        service.create(
            testset_path=invalid,
            request_template="{}",
            run_id="invalid-run",
        )
    assert repository.list_runs() == []
